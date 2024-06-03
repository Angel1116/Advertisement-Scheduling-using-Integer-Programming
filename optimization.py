import pandas as pd
from pulp import LpMaximize, LpProblem, LpVariable, lpSum, LpSolverDefault, LpInteger, LpBinary
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from collections import Counter
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties

user_df = pd.read_csv("user_df.csv", index_col=0)
price_df = pd.read_csv("anime_with_price.csv")
genre_df = pd.read_csv("genre_df.csv")
genre_df = genre_df.set_index("anime_id")

# 製作priority_value的dataframe
def get_affordable_anime_prices(used_df, user_row):
    user_ratings = used_df.iloc[user_row]  
    rated_anime = user_ratings[user_ratings > 0]
    rated_anime.index = rated_anime.index.astype(int)
    new_df = price_df[price_df['anime_id'].isin(rated_anime.index)]

    new_df = new_df.set_index('anime_id')[['price']]
    new_df = pd.concat([new_df, rated_anime.rename('preference')], axis=1)
    new_df['preference'] = new_df['preference'].apply(lambda x: 10*x)
    print("----------------------------")
    
    new_df['score'] = new_df['price'] + new_df['preference']
    new_df['score'] = new_df['score'].apply(lambda x: round(x,0))
    new_df.sort_values(by=['score', 'price'], ascending=[False, False], inplace=True, kind='mergesort')

    new_df.reset_index(inplace=True)
    new_df.rename(columns={'index': 'anime_id'}, inplace=True)
    new_df['priority_value'] = range(len(new_df), 0, -1)
    for i in range(1,len(new_df),1):
        if (new_df.at[i,'score'] == new_df.at[i-1,'score']) and (new_df.at[i,'price'] == new_df.at[i-1,'price']):
            new_df.at[i,'priority_value'] = new_df.at[i-1,'priority_value']
            
    return new_df



    
def optimization(new_df,user,output_name):
    all_anime = list(new_df["anime_id"])
    all_priority = list(new_df["priority_value"])

    new_df = new_df.set_index("anime_id")
    merged_df = new_df.join(genre_df, how="inner")
    top_animes_by_genre = {}
    all_top_anime = []
    
    for genre in genre_df.columns:  #對每個類別挑出priority_value最高的前7部動畫
        top_animes = merged_df[merged_df[genre] == 1].sort_values(by="priority_value", ascending=False).head(7).index.tolist()
        top_animes_by_genre[genre] = top_animes
        all_top_anime.extend(top_animes)
    
    
    num_anime = len(all_top_anime)
    num_days = 7
    num_slots = 14
    # 創建一個 num_anime x num_slots x num_days 的矩陣變數，範圍為 0 或 1 (LpBinary)
    x = [[[LpVariable(f"x_{a}_{i}_{j}", 0, 1, LpBinary) for j in range(num_days)] for i in range(num_slots)] for a in range(num_anime)]

    LpSolverDefault.msg = 0
    
    # 定義問題
    prob = LpProblem("Maximize_Ad_Revenue", LpMaximize)
    
    print("solving...")
    
    # 目標函數：最大化總優先級值
    prob += lpSum([all_priority[all_anime.index(all_top_anime[a])] * x[a][i][j] for a in range(num_anime) for i in range(num_slots) for j in range(num_days)])
    
    # 條件：每個時段最多播放一個廣告
    for i in range(num_slots):
        for j in range(num_days):
            prob += lpSum([x[a][i][j] for a in range(num_anime)]) <= 1, f"One_ad_per_slot_{i}_day_{j}"
    
    # 條件：同一天的廣告不能重複
    for j in range(num_days):
        for a in range(num_anime):
            prob += lpSum([x[a][i][j] for i in range(num_slots)]) <= 1, f"Unique_ad_{a}_day_{j}"
    
    # 條件：同個廣告隔一天同個時間不能出現
    for a in range(num_anime):
        for i in range(num_slots):
            for j in range(num_days - 1):
                prob += x[a][i][j] + x[a][i][j + 1] <= 1, f"Non_repeating_ad_{a}_slot_{i}_day_{j}"
    
    # 條件：同類型的廣告不能連續出現
    for a1 in range(0,num_anime,1):
        for a2 in range(a1+1,num_anime,1):
            common_genres = set(genre_df.loc[all_top_anime[a1]][genre_df.loc[all_top_anime[a1]] == 1].index) & \
                            set(genre_df.loc[all_top_anime[a2]][genre_df.loc[all_top_anime[a2]] == 1].index)
            # 表示找到了兩個同類型的廣告，則他們在相鄰的時段下，就不能夠兩個都一起出現
            if common_genres:
                for i in range(num_slots - 1):
                    for j in range(num_days):
                        prob += x[a1][i][j] + x[a2][i + 1][j] <= 1, f"Non_repeating_genre_ad_{a1}_ad_{a2}_slot_{i}_day_{j}"
                        prob += x[a2][i][j] + x[a1][i + 1][j] <= 1, f"Non_repeating_genre_ad_{a2}_ad_{a1}_slot_{i}_day_{j}"

    prob.solve()
    
    # 提取結果並輸出
    schedule = np.zeros((num_slots, num_days), dtype=int)
    schedule_genre = [["" for _ in range(num_days)] for _ in range(num_slots)]    
    total_priority = 0
    for i in range(num_slots):
        for j in range(num_days):
            for a in range(num_anime):
                if x[a][i][j].varValue == 1:
                    anime_id = all_top_anime[a]
                    schedule[i][j] = anime_id
                    genres = genre_df.loc[anime_id][genre_df.loc[anime_id] == 1].index.tolist()[0]
                    schedule_genre[i][j] = genres
                    #total_priority = total_priority + merged_df.loc[anime_id,"score"]
    

    df_schedule = pd.DataFrame(schedule, columns=[f"Day_{j+1}" for j in range(num_days)], index=[f"Slot_{i+1}" for i in range(num_slots)])
    df_schedule_genre = pd.DataFrame(schedule_genre, columns=[f"Day_{j+1}" for j in range(num_days)], index=[f"Slot_{i+1}" for i in range(num_slots)])
    
    with pd.ExcelWriter(output_name, engine='openpyxl') as writer:
        df_schedule.to_excel(writer, sheet_name="Schedule")
        df_schedule_genre.to_excel(writer, sheet_name="Schedule_Genre")
    
    workbook = load_workbook(output_name)
    font = Font(name='Garamond')
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
    
    workbook.save(output_name)
    print(f"the scedule has been ouputted as '{output_name}'")
    
    return schedule,schedule_genre

def predict(new_data,k):
    sample_df = pd.read_csv(new_data, index_col=0)
    
    common_columns = user_df.columns.intersection(sample_df.columns)
    aligned_user_df = user_df[common_columns]
    aligned_sample_df = sample_df[common_columns]
    sample_row = aligned_sample_df.iloc[0]
    
    def cosine_similarity_ignore_zeros(row1, row2):
        mask = (row1 != 0) & (row2 != 0)
        if mask.sum() == 0:
            return 0  
        row1_filtered = row1[mask]
        row2_filtered = row2[mask]
        
        dot_product = np.dot(row1_filtered, row2_filtered)
        norm_row1 = np.linalg.norm(row1_filtered)
        norm_row2 = np.linalg.norm(row2_filtered)
        
        if norm_row1 == 0 or norm_row2 == 0:
            return 0  
        
        return dot_product / (norm_row1 * norm_row2)
    
    similarities = aligned_user_df.apply(lambda row: cosine_similarity_ignore_zeros(sample_row, row), axis=1)
    
    top_similar_indices = similarities.nlargest(k).index
    # for i in top_similar_indices:
    #     mask = (aligned_user_df.loc[i] != 0) & (sample_row != 0)
    #     print(i,mask.sum(),"相似度:",similarities[top_similar_indices][i])  
    nearest_neighbors = user_df.loc[top_similar_indices]
    
    # predict the user's preferences
    nearest_neighbors = nearest_neighbors.replace(0, np.nan)
    def nanmean_with_check(col):
        if col.isna().all():
            return 0  
        return np.nanmean(col)
    predicted_preferences = nearest_neighbors.apply(nanmean_with_check, axis=0)
    predicted_preferences_df = pd.DataFrame(predicted_preferences)
    transposed_predicted_preferences_df = predicted_preferences_df.transpose()
    return transposed_predicted_preferences_df

def accuracy(original_schedule,predicted_schedule):
    predicted_flat = [item for sublist in predicted_schedule for item in sublist]
    original_flat = [item for sublist in original_schedule for item in sublist]
    predicted_counter = Counter(predicted_flat)
    original_counter = Counter(original_flat)
    all_elements = set(predicted_counter.keys()) | set(original_counter.keys())
    total_overlap = sum(min(predicted_counter[element], original_counter[element]) for element in all_elements)
    
    return total_overlap



print("input '1' to schedule for a user in database")
print("input '2' to upload a new data and schedule for a new user")
choose = int(input("Your choice: "))
if choose == 1:
    user = int(input("user id (0-4417): "))
    new_df = get_affordable_anime_prices(user_df,user)
    schedule = optimization(new_df,user,f"schedule_{user}.xlsx") 
elif choose == 2:
    new_data = input("new data name (with .csv or .xlsx): ")
    data_name = new_data.split('.')[0]
    
    sample_df = pd.read_csv(new_data, index_col=0)
    common_columns = user_df.columns.intersection(sample_df.columns)
    aligned_user_df = user_df[common_columns]
    aligned_sample_df = sample_df[common_columns]
    
    new_df = get_affordable_anime_prices(aligned_sample_df,0) 
    original_schedule,original_genre = optimization(new_df,0,f"schedule_{data_name}.xlsx")
    
    sample_row = sample_df.iloc[0]
    non_zero_count = (sample_row != 0).sum()
    print(f"Number of anime rated by the user is: {non_zero_count}")
    print("If you find this number too small and wish to consider the schedules of users with similar preferences, please refer to the following predictions:")
    
    accuracy_anime_list = []
    accuracy_genre_list = []
    for k in range(3,15,1):
        transposed_predicted_preferences_df= predict(new_data,k)
        new_df = get_affordable_anime_prices(transposed_predicted_preferences_df,0) 
        predicted_schedule, predicted_genre = optimization(new_df,0,f"schedule_predicted_{k}_{data_name}.xlsx")
        
        accuracy_value = accuracy(original_schedule,predicted_schedule)
        accuracy_anime_list.append([k,accuracy_value/98])
        print("k=",k, " anime accuracy=", accuracy_value/98)
        
        accuracy_value = accuracy(original_genre,predicted_genre)
        accuracy_genre_list.append([k,accuracy_value/98])
        print("k=",k, " genre accuracy=", accuracy_value/98)
    

    x_anime = [data[0] for data in accuracy_anime_list]
    y_anime = [data[1] for data in accuracy_anime_list]
    x_genre = [data[0] for data in accuracy_genre_list]
    y_genre = [data[1] for data in accuracy_genre_list]
    
    plt.figure(figsize=(10, 6))
    bar_width = 0.35
    x = range(len(x_anime))
    
    plt.bar(x, y_anime, width=bar_width, label='Anime', color='#83c7f1', align='center')
    plt.bar([p + bar_width for p in x], y_genre, width=bar_width, label='Genre', color='#306998', align='center')
    

    plt.title('Similarity Comparison: Anime vs. Genre', fontsize=16, fontname='Garamond')
    plt.xlabel('Number of Nearest Neighbors (k)', fontsize=14, fontname='Garamond')
    plt.ylabel('Similarity Score', fontsize=14, fontname='Garamond')
    plt.xticks([p + bar_width / 2 for p in x], x_anime, fontname='Garamond')
    plt.yticks(fontname='Garamond')
    plt.ylim(0, 1)
    font = FontProperties()
    font.set_family('Garamond')
    plt.legend(prop=font)
    plt.grid(False)
    
    plt.savefig(f"similarity_for_{data_name}_bar.png", dpi=600)
    plt.show()
